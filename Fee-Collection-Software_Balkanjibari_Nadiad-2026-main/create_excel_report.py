import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime

# Read the CSV file we just created
df = pd.read_csv('enrollment_april_19_2026_complete_10-55-08.csv')

# Create Excel file
excel_path = f"enrollment_april_19_2026_complete_{datetime.now().strftime('%H-%M-%S')}.xlsx"
df.to_excel(excel_path, sheet_name='April 19 Enrollments', index=False)

# Format the Excel file
wb = openpyxl.load_workbook(excel_path)
ws = wb.active
ws.title = 'April 19 Enrollments'

# Define styles
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Style header row
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_alignment
    cell.border = border

# Style data rows
for row_num, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 2):
    for col_idx, cell in enumerate(row, 1):
        cell.border = border
        
        # Alternate row colors
        if row_num % 2 == 0:
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        # Right align numbers
        if col_idx in [1, 7, 8]:  # Sr No, Fee Amount, Amount Paid
            cell.alignment = Alignment(horizontal='right', vertical='center')
        else:
            cell.alignment = left_alignment

# Adjust column widths
ws.column_dimensions['A'].width = 8    # Sr No
ws.column_dimensions['B'].width = 15   # Enrollment ID
ws.column_dimensions['C'].width = 28   # Student Name
ws.column_dimensions['D'].width = 28   # Subject
ws.column_dimensions['E'].width = 15   # Enrollment Date
ws.column_dimensions['F'].width = 14   # Enrollment Time
ws.column_dimensions['G'].width = 12   # Fee Amount
ws.column_dimensions['H'].width = 12   # Amount Paid
ws.column_dimensions['I'].width = 13   # Payment Mode
ws.column_dimensions['J'].width = 15   # Payment Status

# Freeze header row
ws.freeze_panes = 'A2'

# Add summary sheet
ws_summary = wb.create_sheet('Summary', 0)

summary_data = [
    ['ENROLLMENT SUMMARY - APRIL 19, 2026', ''],
    ['', ''],
    ['Total Enrollments', len(df)],
    ['Unique Students', df['Student Name'].nunique()],
    ['Different Subjects', df['Subject'].nunique()],
    ['', ''],
    ['PAYMENT SUMMARY', ''],
    ['Total Amount Collected', f"₹{df['Amount Paid'].sum():,.2f}"],
    ['Online Payments', f"{len(df[df['Payment Mode'] == 'ONLINE'])} | ₹{df[df['Payment Mode'] == 'ONLINE']['Amount Paid'].sum():,.2f}"],
    ['Offline Payments', f"{len(df[df['Payment Mode'] == 'OFFLINE'])} | ₹{df[df['Payment Mode'] == 'OFFLINE']['Amount Paid'].sum():,.2f}"],
    ['', ''],
    ['SUBJECT BREAKDOWN', ''],
]

for row_idx, row_data in enumerate(summary_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_summary.cell(row=row_idx, column=col_idx)
        cell.value = value
        if col_idx == 1:
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        cell.border = border

# Add subject breakdown
start_row = len(summary_data) + 1
for idx, (subject, count) in enumerate(df['Subject'].value_counts().items(), 0):
    amount = df[df['Subject'] == subject]['Amount Paid'].sum()
    row = start_row + idx
    ws_summary.cell(row=row, column=1).value = subject
    ws_summary.cell(row=row, column=2).value = f"{count} | ₹{amount:,.2f}"
    
    for col in [1, 2]:
        cell = ws_summary.cell(row=row, column=col)
        cell.border = border
        if col == 1:
            cell.alignment = left_alignment
        else:
            cell.alignment = left_alignment

ws_summary.column_dimensions['A'].width = 35
ws_summary.column_dimensions['B'].width = 25

# Save the workbook
wb.save(excel_path)

print(f"✅ Excel report created: {excel_path}")
print(f"📊 Total Records: {len(df)}")
print(f"💰 Total Amount: ₹{df['Amount Paid'].sum():,.2f}")
